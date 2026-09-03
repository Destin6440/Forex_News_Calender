from __future__ import annotations
import csv,json
from datetime import datetime,timezone
from pathlib import Path

def safe_cell(v):
    if isinstance(v,str) and v.startswith(("=","+","-","@")): return "'"+v
    return "" if v is None else v
def export_csv(path,results,scope="all_events"):
    path=Path(path); headers=["date_et","matched","rules","event_key","source_event_id","currency","impact_color","event_name","time_et","raw_time","actual","forecast","previous","source_type"]
    with path.open("w",newline="",encoding="utf-8-sig") as f:
        w=csv.DictWriter(f,headers); w.writeheader()
        for result in results:
            matched={k for keys in result.matches.values() for k in keys}
            events=result.events if scope!="matching_dates" else []
            if scope=="matched_events": events=[e for e in events if e.event_key in matched]
            if scope=="matching_dates": w.writerow({"date_et":result.date_et})
            for e in events: w.writerow({k:safe_cell(v) for k,v in {"date_et":result.date_et,"matched":e.event_key in matched,"rules":", ".join(r for r,ks in result.matches.items() if e.event_key in ks),**e.__dict__}.items() if k in headers})
    return path
def export_xlsx(path,results,definition,metadata,scope="all_events"):
    from openpyxl import Workbook
    wb=Workbook(); matches=wb.active; matches.title="Matches"; matches.append(["Date","Matched events","Total events"])
    for r in results: matches.append([r.date_et,r.matched_event_count,len(r.events)])
    events=wb.create_sheet("Events"); fields=["date_et","matched","rules","event_key","source_event_id","currency","impact_color","event_name","time_et","raw_time","actual","forecast","previous","source_type"]; events.append(fields)
    for r in results:
        matched={k for values in r.matches.values() for k in values}
        selected=[] if scope=="matching_dates" else r.events
        if scope=="matched_events":selected=[e for e in selected if e.event_key in matched]
        for e in selected:
            row={"date_et":r.date_et,"matched":e.event_key in matched,"rules":", ".join(rule for rule,keys in r.matches.items() if e.event_key in keys),**e.__dict__}
            events.append([safe_cell(row.get(x)) for x in fields])
    sheet=wb.create_sheet("Search Definition"); sheet.append(["Search name",safe_cell(definition.name)]); sheet.append(["Generated UTC",datetime.now(timezone.utc).isoformat()])
    for k,v in metadata.items(): sheet.append([safe_cell(k),safe_cell(str(v))])
    sheet.append(["Definition JSON",safe_cell(json.dumps(definition.to_dict(),sort_keys=True))]); wb.save(path); return Path(path)
