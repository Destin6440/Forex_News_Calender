"""Atomic, reconciled yearly Excel exports of the canonical calendar."""
from __future__ import annotations

import os
import re
import tempfile
from collections import Counter
from datetime import date, time
from pathlib import Path

WORKBOOK_PATTERN = re.compile(r"^Forex_Factory_News_(\d{4})\.xlsx$")
EVENT_HEADERS = [
    "Event Key", "Date ET", "Time ET", "All Day", "Currency", "Impact Color",
    "Impact Level", "Event Name", "Actual", "Forecast", "Previous", "Datetime ET",
    "Datetime UTC", "Canonical Source Type", "Canonical Source Period",
    "Canonical Source URL", "Source Event ID", "First Seen UTC", "Last Seen UTC",
    "Scraped UTC", "Raw Impact", "Raw Date", "Raw Time",
]
SOURCE_HEADERS = [
    "Event Key", "Provenance Key", "Source Type", "Source Event ID", "Source URL",
    "Source Period", "First Seen UTC", "Last Seen UTC",
]
IMPACT_FILLS = {"red": "F4CCCC", "orange": "FCE5CD", "yellow": "FFF2CC", "gray": "D9D9D9"}


def _openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError as exc:
        raise RuntimeError("Yearly Excel export requires openpyxl; install requirements.txt") from exc


def _typed_date(value):
    return date.fromisoformat(value) if value else None


def _typed_time(value):
    return time.fromisoformat(value) if value else None


def _style_table(sheet, name: str, columns: int, rows: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(columns)}{max(rows, 1)}"
    header_fill = PatternFill("solid", fgColor="17365D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="top")
    if rows >= 2:
        table = Table(displayName=name, ref=f"A1:{get_column_letter(columns)}{rows}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        sheet.add_table(table)


def _widths(sheet, wrap_columns: set[int]) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    for index, cells in enumerate(sheet.columns, 1):
        longest = max((len(str(cell.value)) for cell in cells if cell.value is not None), default=0)
        cap = 55 if index in wrap_columns else 28
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 10), cap)
        if index in wrap_columns:
            for cell in cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _build_workbook(path: Path, year: str, events: list[dict], sources: list[dict], manifest: dict) -> None:
    openpyxl = _openpyxl()
    from openpyxl.styles import Font, PatternFill

    workbook = openpyxl.Workbook()
    event_sheet = workbook.active
    event_sheet.title = "Events"
    event_sheet.append(EVENT_HEADERS)
    for event in events:
        event_sheet.append([
            event["event_key"], _typed_date(event["date_et"]), _typed_time(event["time_et"]),
            bool(event["all_day"]), event["currency"], event["impact_color"], event["impact_level"],
            event["event_name"], event["actual"], event["forecast"], event["previous"],
            event["datetime_et"], event["datetime_utc"], event["source_type"],
            event["source_period"], event["source_url"], event["source_event_id"],
            event["first_seen_at"], event["last_seen_at"], event["scraped_at"],
            event["raw_impact"], event["raw_date"], event["raw_time"],
        ])
    for row in range(2, event_sheet.max_row + 1):
        event_sheet.cell(row, 2).number_format = "yyyy-mm-dd"
        event_sheet.cell(row, 3).number_format = "h:mm AM/PM"
        color = str(event_sheet.cell(row, 6).value or "").lower()
        if color in IMPACT_FILLS:
            event_sheet.cell(row, 6).fill = PatternFill("solid", fgColor=IMPACT_FILLS[color])
    _style_table(event_sheet, f"Events{year}", len(EVENT_HEADERS), event_sheet.max_row)
    _widths(event_sheet, {8, 16})

    source_sheet = workbook.create_sheet("Sources")
    source_sheet.append(SOURCE_HEADERS)
    for source in sources:
        source_sheet.append([source[key] for key in (
            "event_key", "provenance_key", "source_type", "source_event_id", "source_url",
            "source_period", "first_seen_at", "last_seen_at")])
    _style_table(source_sheet, f"Sources{year}", len(SOURCE_HEADERS), source_sheet.max_row)
    _widths(source_sheet, {5})

    summary = workbook.create_sheet("Summary")
    summary.append(["Dataset Summary", "Value"])
    summary_rows = [
        ("Year", int(year)), ("Total canonical events", len(events)),
        ("Earliest event date", _typed_date(events[0]["date_et"])),
        ("Latest event date", _typed_date(events[-1]["date_et"])),
        ("Dataset validation status", "PASS"),
        ("Dataset manifest generation time", manifest.get("generated_at")),
        ("Last successful synchronization time", manifest.get("last_successful_synchronization")),
    ]
    for row in summary_rows:
        summary.append(row)
    sections = (
        ("Counts by month", Counter(e["date_et"][:7] for e in events)),
        ("Counts by currency", Counter(e["currency"] for e in events)),
        ("Counts by impact color", Counter(e["impact_color"] for e in events)),
        ("Counts by source type", Counter(s["source_type"] for s in sources)),
    )
    for title, counts in sections:
        summary.append([]); summary.append([title, "Count"])
        for label, count in sorted(counts.items()): summary.append([label, count])
    summary.freeze_panes = "A2"
    summary.column_dimensions["A"].width = 42; summary.column_dimensions["B"].width = 30
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor="17365D"); cell.font = Font(color="FFFFFF", bold=True)
    summary["B4"].number_format = summary["B5"].number_format = "yyyy-mm-dd"
    workbook.save(path)


def _verify(path: Path, expected_events: int, expected_sources: int) -> None:
    openpyxl = _openpyxl()
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != ["Events", "Sources", "Summary"]:
            raise RuntimeError(f"invalid worksheet structure in {path.name}")
        actual_events = workbook["Events"].max_row - 1
        actual_sources = workbook["Sources"].max_row - 1
        if (actual_events, actual_sources) != (expected_events, expected_sources):
            raise RuntimeError(
                f"workbook reconciliation failed for {path.name}: events {actual_events}/{expected_events}, "
                f"sources {actual_sources}/{expected_sources}"
            )
    finally:
        workbook.close()


def export_yearly_excel(db, output_dir: Path | str, manifest: dict) -> list[Path]:
    """Stage, validate, reconcile, then atomically replace only yearly workbooks."""
    destination = Path(output_dir).expanduser()
    destination.mkdir(parents=True, exist_ok=True)
    events_by_year: dict[str, list[dict]] = {}
    for event in db.rows(): events_by_year.setdefault(event["date_et"][:4], []).append(event)
    sources_by_year: dict[str, list[dict]] = {}
    for source in db.connection.execute(
        "SELECT s.* FROM event_sources s JOIN events e ON e.event_key=s.event_key "
        "ORDER BY e.date_et,e.time_et,e.currency,e.event_name_normalized,e.event_key,s.source_type,s.provenance_key"
    ):
        event_year = db.connection.execute("SELECT substr(date_et,1,4) FROM events WHERE event_key=?", (source["event_key"],)).fetchone()[0]
        sources_by_year.setdefault(event_year, []).append(dict(source))

    with tempfile.TemporaryDirectory(prefix="ff-excel-stage-", dir=destination.parent) as stage_name:
        stage = Path(stage_name)
        staged = []
        for year, events in sorted(events_by_year.items()):
            path = stage / f"Forex_Factory_News_{year}.xlsx"
            sources = sorted(sources_by_year.get(year, []), key=lambda r: (r["event_key"], r["source_type"], r["provenance_key"]))
            _build_workbook(path, year, events, sources, manifest)
            _verify(path, len(events), len(sources))
            staged.append(path)

        with tempfile.TemporaryDirectory(prefix="ff-excel-backup-", dir=destination.parent) as backup_name:
            backup = Path(backup_name)
            existing = [p for p in destination.iterdir() if p.is_file() and WORKBOOK_PATTERN.fullmatch(p.name)]
            backed_up: list[Path] = []
            installed: list[Path] = []
            try:
                for path in existing:
                    backup_path = backup / path.name
                    os.replace(path, backup_path)
                    backed_up.append(backup_path)
                for path in staged:
                    final_path = destination / path.name
                    os.replace(path, final_path)
                    installed.append(final_path)
            except BaseException:
                for path in installed: path.unlink(missing_ok=True)
                for path in backed_up: os.replace(path, destination / path.name)
                raise
    return [destination / f"Forex_Factory_News_{year}.xlsx" for year in sorted(events_by_year)]
