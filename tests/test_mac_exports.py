import csv
from ff_calendar_toolkit.mac_app.exports import export_csv,safe_cell
from ff_calendar_toolkit.mac_app.exports import export_xlsx
from ff_calendar_toolkit.mac_app.filter_schema import SearchDefinition
from ff_calendar_toolkit.mac_app.models import DateResult
from tests.mac_helpers import neutral_event
def test_csv_and_formula_protection(tmp_path):
    assert safe_cell("=unsafe")=="'=unsafe"
    p=export_csv(tmp_path/"x.csv",[DateResult("2025-01-06",[neutral_event("1",name="=unsafe")],{"a":["1"]})])
    assert "'=unsafe" in p.read_text(encoding="utf-8-sig")
def test_xlsx_scopes_sheets_and_formula_protection(tmp_path):
    from openpyxl import load_workbook
    result=DateResult("2025-01-06",[neutral_event("1",name="=unsafe"),neutral_event("2",name="Event Beta")],{"rule alpha":["1"]})
    for scope,expected in [("all_events",3),("matched_events",2),("matching_dates",1)]:
        p=export_xlsx(tmp_path/f"{scope}.xlsx",[result],SearchDefinition(name="+unsafe"),{"database":"@unsafe"},scope);wb=load_workbook(p)
        assert wb.sheetnames==["Matches","Events","Search Definition"] and wb["Events"].max_row==expected
        if expected>1:assert wb["Events"][2][7].value=="'=unsafe"
